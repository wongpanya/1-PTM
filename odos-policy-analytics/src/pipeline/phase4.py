from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from src.analytics.metrics import data_quality_summary, metric_definitions, readiness_scorecard
from src.cleaning.rules import (
    fit_level,
    government_preference_flag,
    normalize_text,
    parse_gpa,
    parse_income_monthly_estimate,
    parse_iso_date,
    standardize_category,
    years_between,
)
from src.governance.audit import append_audit_event
from src.utils.config import PROJECT_ROOT, load_yaml


@dataclass
class PipelineIssue:
    severity: str
    code: str
    odos_uid: str
    field: str
    message: str
    raw_value: str = ""

    def as_dict(self) -> dict[str, str]:
        return {
            "severity": self.severity,
            "code": self.code,
            "odos_uid": self.odos_uid,
            "field": self.field,
            "message": self.message,
            "raw_value": self.raw_value,
        }


def run_phase4_pipeline(config_path: str = "config/phase4_pipeline.yaml") -> dict[str, Any]:
    config = load_yaml(config_path)
    source_path = _resolve_path(config["source"]["default_excel_path"])
    output_dir = _resolve_path(config["outputs"]["directory"])
    output_dir.mkdir(parents=True, exist_ok=True)

    source_hash = _sha256(source_path)
    started_at = datetime.now(timezone.utc).isoformat()
    duplicate_source = _is_duplicate_source(output_dir, source_hash)
    _append_processing_log(
        output_dir,
        "pipeline_started",
        {"source_path": str(source_path), "source_sha256": source_hash, "duplicate_source": duplicate_source},
    )

    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)
    sheet_issues = _validate_sheets(workbook, config)
    if any(issue.severity == "error" for issue in sheet_issues):
        raise ValueError("Required sheets are missing")

    data_sheet = workbook[config["source"]["data_sheet"]]
    dictionary_sheet = workbook[config["source"]["dictionary_sheet"]]
    headers = _read_headers(data_sheet, config)
    column_issues = _validate_required_columns(headers, config)
    dictionary = _read_dictionary(dictionary_sheet, config)

    raw_rows = _read_data_rows(data_sheet, config)
    records, cleaning_stats = _clean_records(raw_rows, config)
    validation_issues = _validate_records(records, dictionary, config)
    all_issues = [issue.as_dict() for issue in [*sheet_issues, *column_issues, *validation_issues]]

    cleaned_df = pd.DataFrame(records)
    issues_df = pd.DataFrame(all_issues)
    definitions = metric_definitions()
    field_report = data_quality_summary(cleaned_df, issues_df, definitions)
    readiness = readiness_scorecard(field_report, definitions)
    before_after = _before_after_report(
        raw_rows,
        cleaned_df,
        all_issues,
        cleaning_stats,
        source_hash,
        duplicate_source,
        workbook.sheetnames,
        headers,
        config,
        field_report,
        readiness,
    )
    quality = _quality_scores(cleaned_df, all_issues, config)
    before_after["quality_scores"] = quality

    cleaned_path = output_dir / config["outputs"]["cleaned_dataset"]
    issues_path = output_dir / config["outputs"]["rejected_dataset"]
    field_report_path = output_dir / config["outputs"]["field_cleaning_report"]
    report_json_path = output_dir / config["outputs"]["before_after_report_json"]
    report_md_path = output_dir / config["outputs"]["before_after_report_md"]

    cleaned_df.to_csv(cleaned_path, index=False, encoding="utf-8-sig")
    issues_df.to_csv(issues_path, index=False, encoding="utf-8-sig")
    field_report.to_csv(field_report_path, index=False, encoding="utf-8-sig")
    report_json_path.write_text(json.dumps(before_after, ensure_ascii=False, indent=2), encoding="utf-8")
    report_md_path.write_text(_report_markdown(before_after), encoding="utf-8")
    _write_import_manifest(output_dir, source_hash, len(raw_rows), len(cleaned_df), duplicate_source)

    completed_at = datetime.now(timezone.utc).isoformat()
    _append_processing_log(
        output_dir,
        "pipeline_completed",
        {
            "completed_at": completed_at,
            "cleaned_rows": len(cleaned_df),
            "issue_count": len(all_issues),
            "duplicate_source": duplicate_source,
            "output_files": [
                str(cleaned_path),
                str(issues_path),
                str(field_report_path),
                str(report_json_path),
                str(report_md_path),
            ],
        },
    )
    append_audit_event("phase4_pipeline_completed", {"rows": len(cleaned_df), "issues": len(all_issues)})

    return {
        "status": "passed",
        "source_path": str(source_path),
        "source_sha256": source_hash,
        "duplicate_source": duplicate_source,
        "rows_read": len(raw_rows),
        "cleaned_rows": len(cleaned_df),
        "issue_count": len(all_issues),
        "quality_scores": quality,
        "outputs": {
            "cleaned_dataset": str(cleaned_path),
            "validation_issues": str(issues_path),
            "field_cleaning_report": str(field_report_path),
            "before_after_report_json": str(report_json_path),
            "before_after_report_md": str(report_md_path),
            "processing_log": str(output_dir / config["outputs"]["processing_log"]),
        },
    }


def _resolve_path(path: str | Path) -> Path:
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = PROJECT_ROOT / candidate
    return candidate.resolve()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _normalize_header(value: Any) -> str:
    return normalize_text(str(value).replace("\n", " ") if value is not None else "") or ""


def _read_headers(sheet, config: dict[str, Any]) -> list[str]:
    header_row = int(config["source"]["header_row"])
    values = next(sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    return [_normalize_header(value) for value in values]


def _validate_sheets(workbook, config: dict[str, Any]) -> list[PipelineIssue]:
    issues = []
    available = set(workbook.sheetnames)
    for sheet in config["required_sheets"]:
        if sheet not in available:
            issues.append(PipelineIssue("error", "missing_sheet", "", "workbook", f"Missing required sheet: {sheet}"))
    return issues


def _validate_required_columns(headers: list[str], config: dict[str, Any]) -> list[PipelineIssue]:
    joined = set(headers)
    issues = []
    for required in config["required_columns"]:
        if required not in joined:
            issues.append(PipelineIssue("error", "missing_required_excel_column", "", required, "Missing required Excel column"))
    return issues


def _read_dictionary(sheet, config: dict[str, Any]) -> dict[str, set[str]]:
    dictionary = {name: set() for name in config["dictionary_columns"].keys()}
    max_row = sheet.max_row or 1
    for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
        for name, one_based_index in config["dictionary_columns"].items():
            value = row[int(one_based_index) - 1] if int(one_based_index) - 1 < len(row) else None
            cleaned = standardize_category(value)
            if cleaned:
                dictionary[name].add(cleaned)
    return dictionary


def _read_data_rows(sheet, config: dict[str, Any]) -> list[tuple]:
    rows = []
    first_data_row = int(config["source"]["first_data_row"])
    for row in sheet.iter_rows(min_row=first_data_row, values_only=True):
        source_id = row[int(config["column_indexes"]["source_id"]) - 1] if row else None
        if normalize_text(source_id) is None:
            continue
        rows.append(row)
    return rows


def _cell(row: tuple, config: dict[str, Any], field: str):
    index = int(config["column_indexes"][field]) - 1
    return row[index] if index < len(row) else None


def _clean_records(rows: list[tuple], config: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, int]]:
    cleaning_config = load_yaml(config.get("cleaning_rules_path", "config/cleaning_rules.yaml"))
    mappings = cleaning_config.get("category_mappings", {})
    university_aliases = mappings.get("university_aliases", {})
    employer_sector_codes = mappings.get("employer_sector_codes", {})
    stats = {
        "blank_to_null": 0,
        "error_formula_to_null": 0,
        "income_parsed": 0,
        "date_parsed": 0,
        "duration_calculated": 0,
    }
    records = []
    for index, row in enumerate(rows, start=1):
        for field in config["column_indexes"].keys():
            value = _cell(row, config, field)
            if isinstance(value, str) and normalize_text(value) is None:
                if value.strip().upper() in {"#NUM!", "#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NAME?"}:
                    stats["error_formula_to_null"] += 1
                else:
                    stats["blank_to_null"] += 1
        source_id = _cell(row, config, "source_id")
        odos_uid = f"ODOS{int(source_id):05d}" if isinstance(source_id, (int, float)) else f"ODOS{index:05d}"
        start_date = parse_iso_date(_cell(row, config, "study_start_date"))
        graduation_date = parse_iso_date(_cell(row, config, "graduation_expected_date"))
        dropout_date = parse_iso_date(_cell(row, config, "study_end_dropout_date"))
        restart_date = parse_iso_date(_cell(row, config, "restart_thailand_date"))
        work_start_date = parse_iso_date(_cell(row, config, "work_start_date"))
        stats["date_parsed"] += sum(1 for value in [start_date, graduation_date, dropout_date, restart_date, work_start_date] if value)
        duration = years_between(start_date, graduation_date)
        if duration is not None:
            stats["duration_calculated"] += 1
        income = parse_income_monthly_estimate(_cell(row, config, "income_raw"))
        if income is not None:
            stats["income_parsed"] += 1

        employment_type = standardize_category(_cell(row, config, "employment_type"))
        record = {
            "odos_uid": odos_uid,
            "source_id": source_id,
            "cohort": _cell(row, config, "cohort"),
            "round": normalize_text(_cell(row, config, "round")),
            "scholarship_type": normalize_text(_cell(row, config, "scholarship_type")),
            "sex": standardize_category(_cell(row, config, "sex")),
            "birth_year_be": _cell(row, config, "birth_year_be"),
            "province_code": _cell(row, config, "province_code"),
            "economic_corridor": standardize_category(_cell(row, config, "economic_corridor")),
            "region": standardize_category(_cell(row, config, "region")),
            "province": _standardize_province(_cell(row, config, "province")),
            "district": standardize_category(_cell(row, config, "district")),
            "study_start_date": start_date,
            "study_end_dropout_date": dropout_date,
            "graduation_expected_date": graduation_date,
            "restart_thailand_date": restart_date,
            "project_condition_status": standardize_category(_cell(row, config, "project_condition_status")),
            "overall_status": standardize_category(_cell(row, config, "overall_status")),
            "after_m6_status": standardize_category(_cell(row, config, "after_m6_status")),
            "current_status": standardize_category(_cell(row, config, "current_status")),
            "current_continent": standardize_category(_cell(row, config, "current_continent")),
            "current_country": _standardize_country(_cell(row, config, "current_country")),
            "current_field_group": standardize_category(_cell(row, config, "current_field_group")),
            "current_field": standardize_category(_cell(row, config, "current_field")),
            "latest_degree": standardize_category(_cell(row, config, "latest_degree")),
            "gpa_numeric": parse_gpa(_cell(row, config, "gpa_raw")),
            "study_duration_years": duration,
            "analysis_year": _analysis_year(start_date, graduation_date, work_start_date),
            "standardized_university_name": standardize_category(
                _cell(row, config, "current_university"),
                university_aliases,
            ),
            "employment_type": employment_type,
            "employer_sector_code": employer_sector_codes.get(employment_type),
            "employment_detail": standardize_category(_cell(row, config, "employment_detail")),
            "workplace_province_country": standardize_category(_cell(row, config, "workplace_province_country")),
            "work_start_date": work_start_date,
            "income_monthly_est": income,
            "domicile_employment_relation": standardize_category(_cell(row, config, "domicile_employment_relation")),
            "field_job_fit": standardize_category(_cell(row, config, "field_job_fit")),
            "field_job_fit_level": fit_level(_cell(row, config, "field_job_fit")),
            "local_fit": standardize_category(_cell(row, config, "local_fit")),
            "local_fit_level": fit_level(_cell(row, config, "local_fit")),
            "job_satisfaction": standardize_category(_cell(row, config, "job_satisfaction")),
            "job_satisfaction_level": fit_level(_cell(row, config, "job_satisfaction")),
            "wants_government": standardize_category(_cell(row, config, "wants_government")),
            "wants_government_flag": government_preference_flag(_cell(row, config, "wants_government")),
        }
        record.update(_targets(record))
        records.append(record)
    return records, stats


def _targets(record: dict[str, Any]) -> dict[str, int]:
    risk_statuses = {"ลาออก", "พ้นสภาพ", "เกินระยะเวลารับทุน", "สละสิทธิ์"}
    employment_ready = {"ภาคเอกชน", "ภาครัฐ", "รัฐวิสาหกิจ", "ธุรกิจส่วนตัว", "องค์กรเอกชนเพื่อสาธารณประโยชน์"}
    return {
        "target_graduation_success": 1 if record.get("project_condition_status") == "สำเร็จการศึกษา" else 0,
        "target_dropout": 1 if record.get("project_condition_status") == "ลาออก" else 0,
        "target_termination": 1 if record.get("project_condition_status") == "พ้นสภาพ" else 0,
        "target_scholarship_risk": 1 if record.get("project_condition_status") in risk_statuses else 0,
        "target_tracking_risk": 1 if record.get("employment_type") == "อยู่ระหว่างติดตามข้อมูล" or "ไม่พบในฐานข้อมูล" in str(record.get("current_status") or "") else 0,
        "target_employment_ready": 1 if record.get("employment_type") in employment_ready else 0,
        "target_field_mismatch": 1 if record.get("field_job_fit_level") in {0, 1} else 0,
        "target_local_mismatch": 1 if record.get("local_fit_level") in {0, 1} else 0,
    }


def _standardize_province(value):
    return standardize_category(value, {"กรุงเทพ": "กรุงเทพมหานคร", "กรุงเทพฯ": "กรุงเทพมหานคร"})


def _standardize_country(value):
    return standardize_category(value, {"ไทย": "ประเทศไทย"})


def _validate_records(records: list[dict[str, Any]], dictionary: dict[str, set[str]], config: dict[str, Any]) -> list[PipelineIssue]:
    issues: list[PipelineIssue] = []
    seen = set()
    income_min = float(config["validation"]["income_min"])
    income_max = float(config["validation"]["income_max"])
    dictionary_checks = config["validation"]["dictionary_checks"]
    dictionary_aliases = {
        field: set(values)
        for field, values in config["validation"].get("dictionary_aliases", {}).items()
    }
    for record in records:
        uid = record["odos_uid"]
        if uid in seen:
            issues.append(PipelineIssue("error", "duplicate_id", uid, "odos_uid", "Duplicate odos_uid"))
        seen.add(uid)

        _date_order_issue(record, uid, issues, "study_start_date", "graduation_expected_date")
        _date_order_issue(record, uid, issues, "study_start_date", "study_end_dropout_date")
        _date_order_issue(record, uid, issues, "study_start_date", "work_start_date", severity="warning")
        _cross_field_issues(record, uid, issues, config)

        income = record.get("income_monthly_est")
        if income is not None and not (income_min <= float(income) <= income_max):
            issues.append(PipelineIssue("warning", "income_out_of_range", uid, "income_monthly_est", f"Income outside configured range {income_min}-{income_max}", str(income)))

        for field, dictionary_name in dictionary_checks.items():
            value = record.get(field)
            allowed = set(dictionary.get(dictionary_name, set()))
            allowed.update(dictionary_aliases.get(field, set()))
            if value and allowed and value not in allowed:
                issues.append(PipelineIssue("warning", "not_in_data_dictionary", uid, field, f"Value not found in dictionary {dictionary_name}", str(value)))

        for field in config["validation"]["completeness_key_fields"]:
            if record.get(field) in {None, ""}:
                issues.append(PipelineIssue("warning", "missing_key_field", uid, field, "Key field is missing"))
    return issues


def _date_order_issue(record, uid, issues, start_field, end_field, severity="error"):
    start = record.get(start_field)
    end = record.get(end_field)
    if start and end and end < start:
        issues.append(PipelineIssue(severity, "date_order_invalid", uid, end_field, f"{end_field} is before {start_field}", f"{start}>{end}"))


def _cross_field_issues(record, uid, issues, config):
    enabled = set(config["validation"].get("cross_field_rules", []))
    completed = record.get("target_graduation_success") == 1
    employed = record.get("target_employment_ready") == 1
    risk = record.get("target_scholarship_risk") == 1
    tracking = record.get("target_tracking_risk") == 1

    if "completed_without_employment" in enabled and completed and (not record.get("employment_type") or tracking):
        issues.append(
            PipelineIssue(
                "warning",
                "completed_without_employment_followup",
                uid,
                "employment_type",
                "Completed recipient has no confirmed employment outcome",
            )
        )
    if "employed_without_work_start_date" in enabled and employed and not record.get("work_start_date"):
        issues.append(
            PipelineIssue(
                "warning",
                "employed_without_work_start_date",
                uid,
                "work_start_date",
                "Employment is recorded but work start date is missing",
            )
        )
    if "graduated_with_dropout_date" in enabled and completed and record.get("study_end_dropout_date"):
        issues.append(
            PipelineIssue(
                "error",
                "graduated_with_dropout_date",
                uid,
                "study_end_dropout_date",
                "Graduation status conflicts with a dropout/end date",
            )
        )
    if "risk_status_marked_employed" in enabled and risk and employed:
        issues.append(
            PipelineIssue(
                "warning",
                "risk_status_marked_employed",
                uid,
                "employment_type",
                "Scholarship-risk status conflicts with an employment-ready status",
            )
        )


def _quality_scores(df: pd.DataFrame, issues: list[dict[str, str]], config: dict[str, Any]) -> dict[str, float]:
    total = max(len(df), 1)
    key_fields = config["validation"]["completeness_key_fields"]
    completeness = float(df[key_fields].notna().mean().mean() * 100) if set(key_fields).issubset(df.columns) else 0.0
    duplicate_count = int(df["odos_uid"].duplicated().sum()) if "odos_uid" in df else total
    uniqueness = max(0.0, 100 - duplicate_count / total * 100)
    error_count = sum(1 for issue in issues if issue["severity"] == "error")
    warning_count = sum(1 for issue in issues if issue["severity"] == "warning")
    validity = max(0.0, 100 - error_count / total * 100)
    consistency = max(0.0, 100 - warning_count / total * 100)
    return {
        "completeness_score": round(completeness, 2),
        "validity_score": round(validity, 2),
        "uniqueness_score": round(uniqueness, 2),
        "consistency_score": round(consistency, 2),
    }


def _before_after_report(
    raw_rows,
    cleaned_df,
    issues,
    cleaning_stats,
    source_hash,
    duplicate_source,
    sheet_names,
    headers,
    config,
    field_report,
    readiness,
):
    issue_counts = pd.DataFrame(issues).groupby("field").size().sort_values(ascending=False).to_dict() if issues else {}
    return {
        "source_sha256": source_hash,
        "duplicate_source": duplicate_source,
        "raw_rows": len(raw_rows),
        "cleaned_rows": int(len(cleaned_df)),
        "cleaned_columns": int(len(cleaned_df.columns)),
        "cleaning_stats": cleaning_stats,
        "issue_count": len(issues),
        "error_count": sum(1 for issue in issues if issue["severity"] == "error"),
        "warning_count": sum(1 for issue in issues if issue["severity"] == "warning"),
        "issue_count_by_column": {str(k): int(v) for k, v in issue_counts.items()},
        "structure_validation": {
            "available_sheets": list(sheet_names),
            "required_sheets": list(config["required_sheets"]),
            "missing_sheets": sorted(set(config["required_sheets"]).difference(sheet_names)),
            "required_columns": list(config["required_columns"]),
            "missing_columns": sorted(set(config["required_columns"]).difference(headers)),
            "output_datatypes": {column: str(dtype) for column, dtype in cleaned_df.dtypes.items()},
        },
        "readiness_scores": readiness.to_dict("records"),
        "field_cleaning_summary": {
            "ready": int((field_report["readiness_status"] == "พร้อมใช้").sum()),
            "aggregate_only": int((field_report["readiness_status"] == "พร้อมใช้เฉพาะ Aggregate").sum()),
            "needs_cleaning_or_review": int(
                field_report["readiness_status"].isin(
                    ["ต้อง clean หรือทบทวนมาตรฐาน", "ต้องเก็บ/ปรับปรุงข้อมูล", "ทบทวนก่อนใช้งาน"]
                ).sum()
            ),
            "excluded_from_ml_features": int(
                field_report["readiness_status"].isin(["ตัดออกจาก ML feature", "Target: ห้ามใช้เป็น feature"]).sum()
            ),
        },
    }


def _analysis_year(*date_values: str | None) -> int | None:
    """Return the first valid year using the configured analytical precedence."""
    for value in date_values:
        if value:
            try:
                return int(str(value)[:4])
            except (TypeError, ValueError):
                continue
    return None


def _report_markdown(report: dict[str, Any]) -> str:
    lines = [
        "# Phase 4 Before-After Report",
        "",
        f"- raw_rows: {report['raw_rows']}",
        f"- cleaned_rows: {report['cleaned_rows']}",
        f"- cleaned_columns: {report['cleaned_columns']}",
        f"- issue_count: {report['issue_count']}",
        f"- error_count: {report['error_count']}",
        f"- warning_count: {report['warning_count']}",
        "",
        "## Quality Scores",
    ]
    for key, value in report["quality_scores"].items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Readiness Scores"])
    for item in report["readiness_scores"]:
        lines.append(
            f"- {item['use_case']}: {item['readiness_score']} "
            f"({item['ready_fields']}/{item['required_fields']} fields, {item['status']})"
        )
    lines.extend(["", "## Cleaning Stats"])
    for key, value in report["cleaning_stats"].items():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Issue Count by Column"])
    for key, value in list(report["issue_count_by_column"].items())[:50]:
        lines.append(f"- {key}: {value}")
    return "\n".join(lines)


def _append_processing_log(output_dir: Path, event_type: str, payload: dict[str, Any]) -> None:
    config = load_yaml("config/phase4_pipeline.yaml")
    path = output_dir / config["outputs"]["processing_log"]
    event = {
        "event_type": event_type,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "payload": payload,
    }
    with path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")


def _manifest_path(output_dir: Path) -> Path:
    return output_dir / "latest_import_manifest.json"


def _is_duplicate_source(output_dir: Path, source_hash: str) -> bool:
    path = _manifest_path(output_dir)
    if not path.exists():
        return False
    try:
        previous = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return False
    return previous.get("source_sha256") == source_hash


def _write_import_manifest(output_dir: Path, source_hash: str, raw_rows: int, cleaned_rows: int, duplicate_source: bool) -> None:
    path = _manifest_path(output_dir)
    payload = {
        "source_sha256": source_hash,
        "raw_rows": raw_rows,
        "cleaned_rows": cleaned_rows,
        "duplicate_source": duplicate_source,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
