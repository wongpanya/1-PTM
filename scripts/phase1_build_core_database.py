import csv
import json
import re
import sqlite3
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl


ROOT = Path(r"C:\Users\Wongpanya.Nu\Documents\1-PTM")
SOURCE_XLSX = ROOT / "690724 DB_ODOS Students+.xlsx"
OUT_DIR = ROOT / "phase1_outputs"
DATA_DIR = OUT_DIR / "data"
REPORT_DIR = OUT_DIR / "reports"
DB_PATH = OUT_DIR / "odos_policy_analytics_prototype.sqlite"

ERROR_TOKENS = {"#NUM!", "#VALUE!", "#REF!", "#DIV/0!", "#N/A", "#NAME?"}


COL = {
    "source_id": 0,
    "integration_area": 1,
    "secondary_education_area": 2,
    "education_region_office": 3,
    "province_code": 4,
    "economic_corridor": 5,
    "region": 6,
    "province": 7,
    "district": 8,
    "sex": 9,
    "cohort": 10,
    "round": 11,
    "scholarship_type": 12,
    "cohort_round": 13,
    "high_school_track": 14,
    "marital_status": 15,
    "religion": 16,
    "previous_scholarship": 17,
    "high_school_name": 18,
    "birth_year_be": 19,
    "registered_subdistrict": 20,
    "registered_district": 21,
    "registered_province": 22,
    "current_subdistrict": 23,
    "current_district": 24,
    "current_province": 25,
    "address_status": 26,
    "initial_status": 27,
    "initial_country": 28,
    "initial_field_group": 29,
    "initial_field": 30,
    "initial_university": 31,
    "initial_city_state": 32,
    "study_start_date": 33,
    "study_end_dropout_date": 34,
    "graduation_expected_date": 38,
    "restart_thailand_date": 42,
    "project_condition_status": 43,
    "overall_status": 44,
    "after_m6_status": 45,
    "current_status": 46,
    "current_continent": 47,
    "current_country": 48,
    "current_field_group": 49,
    "current_field": 50,
    "current_university": 51,
    "current_city_province": 52,
    "domestic_admission_result": 53,
    "domestic_admission_doc_no": 54,
    "graduation_condition_measure": 55,
    "program_funding_years": 56,
    "funding_period_academic_year": 57,
    "remaining_funding_years": 58,
    "contract_no": 59,
    "contract_signed_date": 60,
    "contract_status": 61,
    "latest_degree": 62,
    "gpa_raw": 63,
    "ocs_report_date": 64,
    "moe_report_date": 68,
    "masters_country": 69,
    "masters_field_group": 70,
    "masters_field": 71,
    "masters_university": 72,
    "masters_graduation_date": 73,
    "doctoral_country": 74,
    "doctoral_field_group": 75,
    "doctoral_field": 76,
    "doctoral_university": 77,
    "doctoral_graduation_date": 78,
    "employment_type": 83,
    "employment_detail": 84,
    "job_title": 85,
    "employment_evidence_status": 86,
    "employment_cert_doc_no": 87,
    "employer": 88,
    "workplace_no": 89,
    "workplace_moo": 90,
    "workplace_subdistrict": 91,
    "workplace_district": 92,
    "workplace_province_country": 93,
    "workplace_postcode": 94,
    "workplace_phone": 95,
    "work_start_date": 96,
    "income_raw": 97,
    "welfare": 98,
    "employment_history": 99,
    "workplace_address": 100,
    "domicile_employment_relation": 101,
    "field_job_fit": 102,
    "local_fit": 103,
    "job_satisfaction": 104,
    "wants_government": 105,
    "government_reason": 106,
    "desired_government_position": 107,
    "wants_moe": 108,
    "desired_agency_other": 109,
    "other_comments": 110,
    "officer_notes": 111,
}

PII_OR_SENSITIVE = {
    "contact_prefix",
    "contact_name",
    "contact_relationship",
    "contact_phone",
    "contract_no",
    "employment_cert_doc_no",
    "workplace_no",
    "workplace_moo",
    "workplace_subdistrict",
    "workplace_district",
    "workplace_postcode",
    "workplace_phone",
    "workplace_address",
    "other_comments",
    "officer_notes",
}


def ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)


def normalize_header(value, index):
    if value is None:
        return f"COL{index + 1}"
    return re.sub(r"\s+", " ", str(value).replace("\n", " ")).strip()


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = re.sub(r"\s+", " ", value).strip()
        if not text:
            return None
        if text.upper() in ERROR_TOKENS:
            return None
        return text
    return value


def as_iso_date(value):
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def parse_float(value):
    value = clean_value(value)
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).replace(",", ".")
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    number = float(match.group(0))
    return number


def parse_gpa(value):
    number = parse_float(value)
    if number is None:
        return None
    if 0 <= number <= 4:
        return round(number, 3)
    if 0 <= number <= 100:
        return round(number / 25, 3)
    return None


def parse_income(value):
    value = clean_value(value)
    if value is None:
        return None
    text = str(value).replace(",", "").replace("บาท", "").strip()
    numbers = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", text)]
    if not numbers:
        return None
    if "ขึ้นไป" in str(value):
        estimate = numbers[0]
    elif len(numbers) >= 2:
        estimate = (numbers[0] + numbers[1]) / 2
    else:
        estimate = numbers[0]
    if 0 <= estimate < 1_000_000:
        return round(estimate, 2)
    return None


def duration_years(start, end):
    if not isinstance(start, datetime) or not isinstance(end, datetime):
        return None
    if end < start:
        return None
    return round((end - start).days / 365.25, 2)


def fit_level(value):
    value = clean_value(value)
    if value is None:
        return None
    text = str(value).lower()
    if text in {"n/a", "n/a ", "-", "เสียชีวิต"}:
        return None
    if "ไม่สอดคล้อง" in text:
        return 0
    if "น้อย" in text:
        return 1
    if "ปานกลาง" in text:
        return 2
    if "มาก" in text:
        return 3
    return None


def yes_no_government(value):
    value = clean_value(value)
    if value is None:
        return None
    text = str(value).strip()
    if text.startswith("ต้องการ"):
        return 1
    if text.startswith("ไม่ต้องการ"):
        return 0
    return None


def get(row, key):
    return clean_value(row[COL[key]]) if COL[key] < len(row) else None


def read_source():
    wb = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    ws = wb["DB_Students"]
    headers = [normalize_header(v, i) for i, v in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))]
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if clean_value(row[0]) is not None]
    return wb, headers, rows


def build_records(rows):
    students, education, employment, core = [], [], [], []
    for idx, row in enumerate(rows, start=1):
        source_id = get(row, "source_id")
        odos_uid = f"ODOS{int(source_id):05d}" if isinstance(source_id, (int, float)) else f"ODOS{idx:05d}"

        student = {
            "odos_uid": odos_uid,
            "source_id": source_id,
            "cohort": get(row, "cohort"),
            "round": get(row, "round"),
            "scholarship_type": get(row, "scholarship_type"),
            "cohort_round": get(row, "cohort_round"),
            "sex": get(row, "sex"),
            "birth_year_be": get(row, "birth_year_be"),
            "high_school_track": get(row, "high_school_track"),
            "high_school_name": get(row, "high_school_name"),
            "region": get(row, "region"),
            "province_code": get(row, "province_code"),
            "province": get(row, "province"),
            "district": get(row, "district"),
            "economic_corridor": get(row, "economic_corridor"),
        }

        study_duration = duration_years(get(row, "study_start_date"), get(row, "graduation_expected_date"))
        edu = {
            "odos_uid": odos_uid,
            "initial_status": get(row, "initial_status"),
            "initial_country": get(row, "initial_country"),
            "initial_field_group": get(row, "initial_field_group"),
            "initial_field": get(row, "initial_field"),
            "initial_university": get(row, "initial_university"),
            "study_start_date": as_iso_date(get(row, "study_start_date")),
            "study_end_dropout_date": as_iso_date(get(row, "study_end_dropout_date")),
            "graduation_expected_date": as_iso_date(get(row, "graduation_expected_date")),
            "restart_thailand_date": as_iso_date(get(row, "restart_thailand_date")),
            "project_condition_status": get(row, "project_condition_status"),
            "overall_status": get(row, "overall_status"),
            "after_m6_status": get(row, "after_m6_status"),
            "current_status": get(row, "current_status"),
            "current_continent": get(row, "current_continent"),
            "current_country": get(row, "current_country"),
            "current_field_group": get(row, "current_field_group"),
            "current_field": get(row, "current_field"),
            "current_university": get(row, "current_university"),
            "current_city_province": get(row, "current_city_province"),
            "program_funding_years": get(row, "program_funding_years"),
            "funding_period_academic_year": get(row, "funding_period_academic_year"),
            "contract_status": get(row, "contract_status"),
            "latest_degree": get(row, "latest_degree"),
            "gpa_raw": get(row, "gpa_raw"),
            "gpa_numeric": parse_gpa(get(row, "gpa_raw")),
            "study_duration_years": study_duration,
            "ocs_report_date": as_iso_date(get(row, "ocs_report_date")),
            "moe_report_date": as_iso_date(get(row, "moe_report_date")),
            "masters_country": get(row, "masters_country"),
            "masters_field_group": get(row, "masters_field_group"),
            "masters_field": get(row, "masters_field"),
            "masters_university": get(row, "masters_university"),
            "doctoral_country": get(row, "doctoral_country"),
            "doctoral_field_group": get(row, "doctoral_field_group"),
            "doctoral_field": get(row, "doctoral_field"),
            "doctoral_university": get(row, "doctoral_university"),
        }

        emp = {
            "odos_uid": odos_uid,
            "employment_type": get(row, "employment_type"),
            "employment_detail": get(row, "employment_detail"),
            "job_title": get(row, "job_title"),
            "employment_evidence_status": get(row, "employment_evidence_status"),
            "employer": get(row, "employer"),
            "workplace_province_country": get(row, "workplace_province_country"),
            "work_start_date": as_iso_date(get(row, "work_start_date")),
            "income_raw": get(row, "income_raw"),
            "income_monthly_est": parse_income(get(row, "income_raw")),
            "welfare": get(row, "welfare"),
            "domicile_employment_relation": get(row, "domicile_employment_relation"),
            "field_job_fit": get(row, "field_job_fit"),
            "field_job_fit_level": fit_level(get(row, "field_job_fit")),
            "local_fit": get(row, "local_fit"),
            "local_fit_level": fit_level(get(row, "local_fit")),
            "job_satisfaction": get(row, "job_satisfaction"),
            "job_satisfaction_level": fit_level(get(row, "job_satisfaction")),
            "wants_government": get(row, "wants_government"),
            "wants_government_flag": yes_no_government(get(row, "wants_government")),
            "desired_government_position": get(row, "desired_government_position"),
            "wants_moe": get(row, "wants_moe"),
            "desired_agency_other": get(row, "desired_agency_other"),
        }

        core_row = {}
        core_row.update(student)
        core_row.update({k: v for k, v in edu.items() if k != "odos_uid"})
        core_row.update({k: v for k, v in emp.items() if k != "odos_uid"})

        students.append(student)
        education.append(edu)
        employment.append(emp)
        core.append(core_row)
    return students, education, employment, core


def write_csv(path, rows):
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def sqlite_type(value):
    if isinstance(value, int):
        return "INTEGER"
    if isinstance(value, float):
        return "REAL"
    return "TEXT"


def create_table(conn, table_name, rows, primary_key=None):
    if not rows:
        return
    columns = list(rows[0].keys())
    samples = {c: next((r[c] for r in rows if r.get(c) is not None), "") for c in columns}
    column_defs = []
    for column in columns:
        col_type = sqlite_type(samples[column])
        pk = " PRIMARY KEY" if column == primary_key else ""
        column_defs.append(f'"{column}" {col_type}{pk}')
    conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
    conn.execute(f'CREATE TABLE "{table_name}" ({", ".join(column_defs)})')
    placeholders = ", ".join(["?"] * len(columns))
    quoted = ", ".join([f'"{c}"' for c in columns])
    conn.executemany(
        f'INSERT INTO "{table_name}" ({quoted}) VALUES ({placeholders})',
        [[r.get(c) for c in columns] for r in rows],
    )


def build_dictionary(headers, rows):
    selected_reverse = {v: k for k, v in COL.items()}
    dictionary = []
    for i, header in enumerate(headers):
        values = [clean_value(row[i]) if i < len(row) else None for row in rows]
        non_empty = sum(v is not None for v in values)
        errors = sum(isinstance(row[i], str) and str(row[i]).strip().upper() in ERROR_TOKENS for row in rows if i < len(row))
        top_values = Counter(str(v) for v in values if v is not None).most_common(5)
        mapped = selected_reverse.get(i, "")
        if i <= 26:
            category = "student_profile_and_location"
        elif i <= 63:
            category = "education_and_scholarship"
        elif i <= 78:
            category = "postgraduate_tracking"
        elif i <= 82:
            category = "contact_pii_excluded"
        else:
            category = "employment_and_followup"
        dictionary.append({
            "original_col_no": i + 1,
            "original_header": header,
            "mapped_field": mapped,
            "category": category,
            "include_in_prototype": "no" if mapped in PII_OR_SENSITIVE or category == "contact_pii_excluded" else ("yes" if mapped else "review"),
            "non_empty_count": non_empty,
            "missing_count": len(rows) - non_empty,
            "completeness_pct": round(non_empty / len(rows) * 100, 2) if rows else 0,
            "error_like_count": errors,
            "top_values_sample": " | ".join(f"{v} ({c})" for v, c in top_values),
        })
    return dictionary


def build_quality_summary(headers, rows, students, education, employment):
    def count_values(records, field):
        return Counter(str(r.get(field)) for r in records if r.get(field) is not None)

    completed = sum(r.get("project_condition_status") == "สำเร็จการศึกษา" for r in education)
    income_available = sum(r.get("income_monthly_est") is not None for r in employment)
    gpa_available = sum(r.get("gpa_numeric") is not None for r in education)
    field_fit_available = sum(r.get("field_job_fit_level") is not None for r in employment)
    local_fit_available = sum(r.get("local_fit_level") is not None for r in employment)

    metrics = [
        ("source_file", SOURCE_XLSX.name),
        ("source_rows_with_id", len(rows)),
        ("source_columns", len(headers)),
        ("unique_source_id", len({r["source_id"] for r in students})),
        ("cohorts", ", ".join(f"{k}: {v}" for k, v in sorted(count_values(students, "cohort").items()))),
        ("completion_count", completed),
        ("completion_rate_pct", round(completed / len(rows) * 100, 2)),
        ("province_distinct", len(count_values(students, "province"))),
        ("district_distinct", len(count_values(students, "district"))),
        ("current_country_distinct", len(count_values(education, "current_country"))),
        ("current_field_group_distinct", len(count_values(education, "current_field_group"))),
        ("employment_type_distinct", len(count_values(employment, "employment_type"))),
        ("gpa_numeric_available", gpa_available),
        ("income_monthly_est_available", income_available),
        ("field_job_fit_available", field_fit_available),
        ("local_fit_available", local_fit_available),
    ]
    return [{"metric": k, "value": v} for k, v in metrics]


def write_report(quality, dictionary, students, education, employment):
    quality_map = {row["metric"]: row["value"] for row in quality}
    dict_sorted = sorted(dictionary, key=lambda r: r["completeness_pct"])
    report = [
        "# Phase 1 Data Quality Report",
        "",
        "## Scope",
        "รายงานนี้สร้างจากไฟล์ `690724 DB_ODOS Students+.xlsx` โดยอ่านชีต `DB_Students` และใช้เฉพาะแถวที่มี `ID` เพื่อสร้างฐานข้อมูลกลางแบบ Prototype",
        "",
        "## Core Metrics",
    ]
    for row in quality:
        report.append(f"- {row['metric']}: {row['value']}")

    report.extend([
        "",
        "## Prototype Tables",
        "- `students`: ข้อมูลผู้รับทุนเชิงพื้นที่และข้อมูลพื้นฐานที่ไม่ใช้ข้อมูลติดต่อส่วนบุคคล",
        "- `education`: ข้อมูลสถานะการศึกษา ทุน ประเทศ สาขา วุฒิ GPA และระยะเวลาศึกษา",
        "- `employment`: ข้อมูลอาชีพ รายได้โดยประมาณ ความสอดคล้องงานกับสาขา/ท้องถิ่น และความต้องการรับราชการ",
        "- `annual_external_indicators_template`: โครงสร้างสำหรับกรอกข้อมูลเสริมรายปีในระยะถัดไป",
        "",
        "## Data Readiness Notes",
        f"- ข้อมูลผู้รับทุนที่ใช้ได้: {quality_map['source_rows_with_id']} records",
        f"- อัตราสำเร็จการศึกษาจากสถานะโครงการ: {quality_map['completion_rate_pct']}%",
        f"- GPA numeric ใช้ได้: {quality_map['gpa_numeric_available']} records",
        f"- รายได้แปลงเป็นค่าประมาณรายเดือนได้: {quality_map['income_monthly_est_available']} records",
        "- คอลัมน์วันที่แยกปี/เดือน/วันบางส่วนมีค่า error-like จำนวนมาก จึงควรใช้คอลัมน์วันที่เต็มในการวิเคราะห์เวลา",
        "- Prototype นี้ไม่ export ข้อมูลติดต่อส่วนบุคคล เช่น ชื่อผู้ติดต่อ โทรศัพท์ เลขที่สัญญา และที่อยู่ละเอียด",
        "",
        "## Lowest Completeness Fields",
    ])
    for row in dict_sorted[:15]:
        report.append(f"- Col {row['original_col_no']} `{row['original_header']}`: {row['completeness_pct']}%")

    (REPORT_DIR / "phase1_data_quality_report.md").write_text("\n".join(report), encoding="utf-8")


def write_external_indicator_template():
    rows = [
        {
            "indicator_id": "COST_COUNTRY_YYYY_001",
            "fiscal_year": 2569,
            "indicator_group": "cost",
            "indicator_name": "average_scholarship_cost_per_student",
            "level": "country",
            "country": "ตัวอย่างประเทศ",
            "province": "",
            "field_group": "",
            "industry": "",
            "unit": "THB/year",
            "value": "",
            "source_name": "",
            "source_url": "",
            "updated_at": "",
            "confidence_level": "draft",
            "notes": "ใช้สำหรับ ROI/SROI ในระยะถัดไป",
        },
        {
            "indicator_id": "LABOR_FIELD_YYYY_001",
            "fiscal_year": 2569,
            "indicator_group": "labor_market",
            "indicator_name": "future_workforce_demand_index",
            "level": "field_group",
            "country": "",
            "province": "",
            "field_group": "กลุ่มวิศวกรรม",
            "industry": "",
            "unit": "index",
            "value": "",
            "source_name": "",
            "source_url": "",
            "updated_at": "",
            "confidence_level": "draft",
            "notes": "ใช้สำหรับ Future Workforce Demand และ Field Recommendation",
        },
        {
            "indicator_id": "AREA_PROVINCE_YYYY_001",
            "fiscal_year": 2569,
            "indicator_group": "provincial_indicator",
            "indicator_name": "provincial_priority_score",
            "level": "province",
            "country": "ประเทศไทย",
            "province": "ตัวอย่างจังหวัด",
            "field_group": "",
            "industry": "",
            "unit": "score",
            "value": "",
            "source_name": "",
            "source_url": "",
            "updated_at": "",
            "confidence_level": "draft",
            "notes": "ใช้สำหรับ Area-based Allocation",
        },
    ]
    write_csv(DATA_DIR / "annual_external_indicators_template.csv", rows)
    return rows


def main():
    ensure_dirs()
    wb, headers, rows = read_source()
    students, education, employment, core = build_records(rows)
    dictionary = build_dictionary(headers, rows)
    quality = build_quality_summary(headers, rows, students, education, employment)
    indicators = write_external_indicator_template()

    write_csv(DATA_DIR / "students.csv", students)
    write_csv(DATA_DIR / "education.csv", education)
    write_csv(DATA_DIR / "employment.csv", employment)
    write_csv(DATA_DIR / "core_scholarship_dataset.csv", core)
    write_csv(DATA_DIR / "data_dictionary.csv", dictionary)
    write_csv(DATA_DIR / "data_quality_summary.csv", quality)

    if DB_PATH.exists():
        DB_PATH.unlink()
    conn = sqlite3.connect(DB_PATH)
    create_table(conn, "students", students, primary_key="odos_uid")
    create_table(conn, "education", education, primary_key="odos_uid")
    create_table(conn, "employment", employment, primary_key="odos_uid")
    create_table(conn, "core_scholarship_dataset", core, primary_key="odos_uid")
    create_table(conn, "data_dictionary", dictionary)
    create_table(conn, "data_quality_summary", quality)
    create_table(conn, "annual_external_indicators_template", indicators)
    conn.commit()
    conn.close()

    schema = {
        "database": str(DB_PATH),
        "source_file": str(SOURCE_XLSX),
        "tables": {
            "students": list(students[0].keys()),
            "education": list(education[0].keys()),
            "employment": list(employment[0].keys()),
            "core_scholarship_dataset": list(core[0].keys()),
            "data_dictionary": list(dictionary[0].keys()),
            "data_quality_summary": list(quality[0].keys()),
            "annual_external_indicators_template": list(indicators[0].keys()),
        },
        "privacy_note": "Prototype exports exclude direct contact columns and detailed workplace address fields where possible. Use aggregate views for policy presentation.",
    }
    (REPORT_DIR / "phase1_schema.json").write_text(json.dumps(schema, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_report(quality, dictionary, students, education, employment)

    print(json.dumps({
        "rows": len(rows),
        "outputs": {
            "database": str(DB_PATH),
            "data_dir": str(DATA_DIR),
            "report_dir": str(REPORT_DIR),
        },
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
